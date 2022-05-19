import copy
import openpyxl
from tkinter import *
import tkinter as tk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import functions as f

# Функции для обработки введенных данных
def Select():
    start = s.get()
    end = e.get()
    temp = t.get()
    print(start, end, temp)

    #Обработка данных
    lst = f.Counter(start, end, temp, int(ntr2.get()))

    smt = f.Smt(lst, float(ntr3.get()))

    mnk = f.Mnk(smt)

    l_aprk = f.Aprk(lst)
    s_aprk = f.Aprk(smt)

    # Работа с таблицей
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(len(l_aprk)):
        ws.cell(row=i+1, column=1).value = l_aprk[i]
    for i in range(len(s_aprk)):
        ws.cell(row=i+1, column=2).value = s_aprk[i]
    for i in range(len(mnk)):
        ws.cell(row=i+1, column=3).value = mnk[i]
    wb.save('table.xlsx')
    wb.close()

    #Работа с графиком
    Generate(lst, smt, mnk, l_aprk, s_aprk)


def Generate(lst, smt, mnk, l_aprk, s_aprk):

    global conv
    if conv:
        conv.get_tk_widget().destroy()

    figure2 = plt.Figure(figsize=(16, 9), dpi=100)
    ax2 = figure2.add_subplot(111)
    conv = FigureCanvasTkAgg(figure2, str5)
    conv.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH)
    #ax2.plot(smt)
    #ax2.plot(lst)
    ax2.plot(mnk)
    ax2.plot(l_aprk)
    ax2.plot(s_aprk)
    return conv


def Close():
    global conv
    if conv:
        conv.get_tk_widget().destroy()
        return conv

# Структура окна
window = Tk()
window.geometry("900x700")
window.title("Аналитика функций")

conv = None

str1 = Frame(window)
str2 = Frame(window)
str3 = Frame(window)
str4 = Frame(window)
str5 = Frame(window)


str1.pack()
str2.pack()
str3.pack()
str4.pack()
str5.pack()

lb2 = Label(str1, text="Введите начало, конец и шаг значений:", padx=5, pady=5)
lb2.pack(side=LEFT)


s = Entry(str1, width=15)
e = Entry(str1, width=15)
t = Entry(str1, width=15)
s.pack(side=LEFT)
e.pack(side=LEFT)
t.pack(side=LEFT)

lb5 = Label(str2, text="Номер функции:", padx=5, pady=5)
lb5.pack(side=LEFT)

lb6 = Label(str3, text="Коэффициент сглаживания:", padx=5, pady=5)
lb6.pack(side=LEFT)

ntr2 = Entry(str2, width=15)
ntr2.pack(side=LEFT)

ntr3 = Entry(str3, width=15)
ntr3.pack(side=LEFT)

btn1 = Button(str4, text="Построить график", command=Select, padx=5, pady=5)
btn1.pack(side=LEFT)

button = Button(str4, text="Отчистить", command=Close, padx=5, pady=5)
button.pack(side=LEFT)

window.mainloop()